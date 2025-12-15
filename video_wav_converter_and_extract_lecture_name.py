import pandas
import tkinter as tk
from tkinter import filedialog
import openpyxl
import subprocess
import os
from pathlib import Path
import re
from natsort import natsorted

def extract_audio_from_video(video_files, output_dir):
    # ffmpeg 경로를 절대 경로로 지정
    ffmpeg_path = r"C:\Users\cufs\Desktop\업무\subtitle\ffmpeg\ffmpeg.exe"  # ffmpeg 경로를 실제 설치된 경로로 수정
    

    for video_file in list(video_files):
        if video_file.endswith(".mp4"):
            output_file = os.path.join(output_dir, f"{os.path.splitext(os.path.basename(video_file))[0]}.wav")

            command = [ffmpeg_path, "-i", video_file, "-vn", "-acodec", "pcm_s16le", "-ar", "16000", "-ac", "1", output_file]
            try:
                subprocess.run(command, check=True)
                print(f"오디오 추출 완료: {output_file}")
            except subprocess.CalledProcessError as e:
                print(f"에러 발생: {e}")
                
                
def extract_filename_xlsx(video_files, output_path):
    """다수의 영상파일명에서 공통적인 강의명을 추출해 엑셀파일로 뽑아냅니다."""
    filenames = []
    for file in video_files:
        filename = re.sub(r'\s+|_', "", Path(file).stem)
        filename = re.sub(r'0(\d)주차', r'\1주차', filename)
        filenames.append(filename)
    
    filenames = natsorted(filenames)
    # print(filenames)

    final_lecture_name = set() # 최종강의명을 넣을 set
    temp_list = filenames[:8] # 파일명 리스트의 앞 8개 항목을 임시리스트에 넣음
    filenames = filenames[8:] # 파일명 리스트에서 앞 8개 항목을 빼고 재정의
    while filenames:
        for i, filename in enumerate(filenames):
            temp_lecture_name = os.path.commonprefix(temp_list) # 임시리스트의 공통접두사
            temp_list.append(filename) # 임시리스트에 현재파일 추가
            new_temp_lecture_name = os.path.commonprefix(temp_list) # 현재파일이 추가된 임시리스트의 공통접두사
            if temp_lecture_name == new_temp_lecture_name: # 그것이 서로 같으면
                continue # 계속 진행
            else: # 같지 않으면 현재파일은 같은 강의가 아닌 새로운 강의가 나오는 경계지점
                final_lecture_name.add(temp_lecture_name) # 현재파일을 추가하기 전 임시리스트의 공통접두사를 최종강의명으로 추가
                temp_list = filenames[i:i+8] # 현재파일부터 8개의 파일을 새 임시리스트로 재정의
                filenames = filenames[i+8:] # 파일명 리스트를 위의 파일을 제외하고 재정의
                break # for 문을 나가기
        else:
            filenames = [] # for 문을 전부 실행한 후 while 문을 나가기 위한 조건 생성
            
            
    if temp_list:
        final_lecture_name.add(os.path.commonprefix(temp_list)) # 혹시 모를 남아있는 공통접두사를 최종강의명에 추가
        
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws['C4'].value = '강의명'
    ws['D4'].value = '보조언어'
    ws['G4'].value = '언어별 인덱스'
    ws['G5'].value = '한국어 : ko'
    ws['G6'].value = '영어 : en'
    ws['G7'].value = '스페인어 : es'
    ws['G8'].value = '일본어 : ja'
    ws['G9'].value = '중국어 : zh'
    ws['G10'].value = '베트남어 : vi'
    ws['G11'].value = '인도네시아어 : id'
    
    
    for i, name in enumerate(final_lecture_name):
        ws[f'C{i+5}'] = name
    

    wb.save(output_path)
    
    
### 실행부 ###    

#Tkinter 창 숨기기(백 그라운드로)
root = tk.Tk()
root.withdraw()

# 예시 사용
video_files = filedialog.askopenfilenames(
            title="mp4 파일 선택 (다중 선택 가능)",
            filetypes=[("mp4 Files", "*.mp4")]
        )
output_directory = filedialog.askdirectory(title="WAV파일을 저장할 경로 선택")

xlsx_file = '강의명 추출'
excel_path = filedialog.asksaveasfilename(
    title="Excel 파일을 저장하세요",
    defaultextension=".xlsx",
    filetypes=[("Excel", "*.xlsx")],
    initialfile=xlsx_file
    )


extract_filename_xlsx(video_files, excel_path)
extract_audio_from_video(video_files, output_directory)